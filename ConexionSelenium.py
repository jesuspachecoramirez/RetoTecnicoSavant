from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.utils import COMMASPACE
from email import encoders
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook

def envio_correo():
    destinatario_ingresado = input("Por favor ingrese su correo electronico para el envio de informacion\n")
    msg = MIMEMultipart()
    msg['From'] = 'pruebatecnicasavant@gmail.com'
    msg['To'] = COMMASPACE.join([destinatario_ingresado])
    msg['Subject'] = 'Archivo de Excel con los datos de los productos'
    archivo_excel = 'pruebaTecnica.xlsx'

    with open(archivo_excel, 'rb') as f:
        archivo_adjunto = MIMEBase('application', 'octet-stream')
        archivo_adjunto.set_payload(f.read())
        encoders.encode_base64(archivo_adjunto)
        archivo_adjunto.add_header('Content-Disposition', 'attachment', filename=archivo_excel)
        msg.attach(archivo_adjunto)
        servidor_smtp = 'smtp.gmail.com'
        puerto_smtp = 587
        usuario = 'pruebatecnicasavant@gmail.com'
        contraseña = 'rnwphebtvujyaxwy'

    with smtplib.SMTP(servidor_smtp, puerto_smtp) as smtp:
        smtp.starttls()
        smtp.login(usuario, contraseña)
        smtp.sendmail(msg['From'], msg['To'], msg.as_string())
    return

def busqueda_mercadolibre():
    product =  input("Ingrese el producto a buscar\n")
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.maximize_window()
    driver.get("https://www.mercadolibre.com.co/")
    time.sleep(3)
    driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div[3]/button[1]").click()
    datos = driver.find_element(By.ID, "cb1-edit")
    datos.send_keys(product)
    datos.send_keys(Keys.ENTER)
    calidad = driver.find_element(By.PARTIAL_LINK_TEXT, "Mejores vendedores")
    hover = ActionChains(driver).move_to_element(calidad)
    hover.click()
    hover.perform()
    time.sleep(3)
    driver.find_element(By.XPATH, "//*[@id='root-app']/div/div[2]/section/div[2]/div[2]/div/div/div[2]/div/div/button/span").click()
    time.sleep(1)
    driver.find_element(By.XPATH, "//*[@id='andes-dropdown-más-relevantes-list-option-price_asc']/div/div/span").click()
    time.sleep(3)
    wb = Workbook()
    hoja = wb.active
    hoja.cell(row=1 , column=1, value="Producto")
    hoja.cell(row=2 , column=1, value=driver.find_element(By.XPATH, "//*[@id='root-app']/div/div[2]/section/ol/li[1]").text)
    hoja.cell(row=3 , column=1 , value=driver.find_element(By.XPATH, "//*[@id='root-app']/div/div[2]/section/ol/li[2]").text)
    hoja.cell(row=4 , column=1 , value=driver.find_element(By.XPATH, "//*[@id='root-app']/div/div[2]/section/ol/li[3]").text)
    hoja.cell(row=1 , column=2, value= "Link")
    hoja.cell(row=2, column=2, value=driver.find_element(By.XPATH, "//*[@id='root-app']/div/div[2]/section/ol/li[1]/div/div/div[1]/a").get_attribute("href"))
    hoja.cell(row=3, column=2, value=driver.find_element(By.XPATH, "//*[@id='root-app']/div/div[2]/section/ol/li[2]/div/div/div[1]/a").get_attribute("href"))
    hoja.cell(row=4, column=2, value=driver.find_element(By.XPATH, "//*[@id='root-app']/div/div[2]/section/ol/li[3]/div/div/div[1]/a").get_attribute("href"))
    wb.save("pruebaTecnica.xlsx")
    driver.close()

    return

def busqueda_amazon():
    product = input("Ingrese el producto a buscar\n")
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.maximize_window()
    driver.get("https://www.amazon.com")
    time.sleep(2)
    datos = driver.find_element(By.ID, "twotabsearchtextbox")
    datos.send_keys(product)
    datos.send_keys(Keys.ENTER)
    calidad = driver.find_element(By.XPATH, "//*[@id='reviewsRefinements']/ul/span[1]")
    calidad.click()
    time.sleep(3)
    driver.find_element(By.XPATH, "//*[@id='a-autoid-0-announce']/span[1]").click()
    time.sleep(1)
    driver.find_element(By.ID, "s-result-sort-select_1").click()
    time.sleep(5)
    wb = Workbook()
    hoja = wb.active
    hoja.cell(row=1 , column=1, value="Producto")
    hoja.cell(row=2 , column=1, value=driver.find_element(By.XPATH, "//*[@id='search']/div[1]/div[1]/div/span[1]/div[1]/div[3]").text)
    hoja.cell(row=3 , column=1 , value=driver.find_element(By.XPATH, "//*[@id='search']/div[1]/div[1]/div/span[1]/div[1]/div[4]").text)
    hoja.cell(row=4 , column=1 , value=driver.find_element(By.XPATH, "//*[@id='search']/div[1]/div[1]/div/span[1]/div[1]/div[5]").text)
    hoja.cell(row=1 , column=2, value= "Link")
    hoja.cell(row=2, column=2, value=driver.find_element(By.XPATH, "//*[@id='search']/div[1]/div[1]/div/span[1]/div[1]/div[3]/div/div/div/div/div[2]/div[1]/h2/a").get_attribute("href"))
    hoja.cell(row=3, column=2, value=driver.find_element(By.XPATH, "//*[@id='search']/div[1]/div[1]/div/span[1]/div[1]/div[4]/div/div/div/div/div[2]/div[1]/h2/a").get_attribute("href"))
    hoja.cell(row=4, column=2, value=driver.find_element(By.XPATH, "//*[@id='search']/div[1]/div[1]/div/span[1]/div[1]/div[5]/div/div/div/div/div[2]/div[1]/h2/a").get_attribute("href"))
    wb.save("pruebaTecnica.xlsx")
    driver.close()

    return

def busqueda_bestbuy():
    product =  input("Enter the product to search\n")
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.maximize_window()
    driver.get("https://www.bestbuy.com/")
    time.sleep(4)
    driver.find_element(By.XPATH, "/html/body/div[2]/div/div/div/div[3]/div[2]/a[2]/img").click()
    time.sleep(3)
    datos = driver.find_element(By.ID, "gh-search-input")
    datos.send_keys(product)
    datos.send_keys(Keys.ENTER)
    time.sleep(2)
    calidad = driver.find_element(By.ID, "customerreviews_facet-5-1")
    hover = ActionChains(driver).move_to_element(calidad).click(calidad)
    hover.perform()
    time.sleep(2)
    precioBajo = Select(driver.find_element(By.ID, "sort-by-select"))
    precioBajo.select_by_value("Price-Low-To-High")
    time.sleep(5)
    wb = Workbook()
    hoja = wb.active
    hoja.cell(row=1, column=1, value="Producto")
    hoja.cell(row=2, column=1, value=driver.find_element(By.XPATH, "//*[@id='main-results']/ol/li[2]").text)
    hoja.cell(row=3, column=1, value=driver.find_element(By.XPATH, "//*[@id='main-results']/ol/li[3]").text)
    hoja.cell(row=4, column=1, value=driver.find_element(By.XPATH, "//*[@id='main-results']/ol/li[4]").text)
    hoja.cell(row=1, column=2, value="Link")
    hoja.cell(row=2, column=2, value=driver.find_element(By.XPATH, "/html/body/div[5]/main/div[12]/div/div/div/div/div/div/div[2]/div[2]/div[5]/div/div[4]/ol/li[2]/div/div/div/div/div/div[2]/h4/a").get_attribute("href"))
    hoja.cell(row=3, column=2, value=driver.find_element(By.XPATH,"/html/body/div[5]/main/div[12]/div/div/div/div/div/div/div[2]/div[2]/div[5]/div/div[4]/ol/li[3]/div/div/div/div/div/div[2]/h4/a").get_attribute("href"))
    hoja.cell(row=4, column=2, value=driver.find_element(By.XPATH, "/html/body/div[5]/main/div[12]/div/div/div/div/div/div/div[2]/div[2]/div[5]/div/div[4]/ol/li[4]/div/div/div/div/div/div[2]/h4/a").get_attribute("href"))
    wb.save("pruebaTecnica.xlsx")
    driver.close()

    return

def busqueda_aliexpress():
    product = input("Ingrese el producto a buscar\n")
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.maximize_window()
    driver.get("https://www.aliexpress.com")
    time.sleep(1)
    datos = driver.find_element(By.ID, "search-key")
    datos.send_keys(product)
    datos.send_keys(Keys.ENTER)
    driver.find_element(By.XPATH, "//*[@id='root']/div[1]/div/div[2]/div/div[2]/div[1]/div/div[3]/div[5]/div/label/span[1]").click()
    time.sleep(3)
    driver.find_element(By.XPATH, "//*[@id='root']/div[1]/div/div[2]/div/div[2]/div[2]/div[1]/div/div/div[3]/div/span[2]").click()
    time.sleep(5)
    wb = Workbook()
    hoja = wb.active
    hoja.cell(row=1, column=1, value="Producto")
    hoja.cell(row=2, column=1, value=driver.find_element(By.XPATH, "//*[@id='root']/div[1]/div/div[2]/div/div[2]/div[3]/a[1]").text)
    hoja.cell(row=3, column=1, value=driver.find_element(By.XPATH, "//*[@id='root']/div[1]/div/div[2]/div/div[2]/div[3]/a[2]").text)
    hoja.cell(row=4, column=1, value=driver.find_element(By.XPATH, "//*[@id='root']/div[1]/div/div[2]/div/div[2]/div[3]/a[3]").text)
    hoja.cell(row=1, column=2, value="Link")
    hoja.cell(row=2, column=2, value=driver.find_element(By.XPATH, "//*[@id='root']/div[1]/div/div[2]/div/div[2]/div[3]/a[1]").get_attribute("href"))
    hoja.cell(row=3, column=2, value=driver.find_element(By.XPATH,"//*[@id='root']/div[1]/div/div[2]/div/div[2]/div[3]/a[2]").get_attribute("href"))
    hoja.cell(row=4, column=2, value=driver.find_element(By.XPATH, "//*[@id='root']/div[1]/div/div[2]/div/div[2]/div[3]/a[3]").get_attribute("href"))
    wb.save("pruebaTecnica.xlsx")
    driver.close()

    return

def menu():

        print("Opciones de Mercado")
        print("1. MercadoLibre")
        print("2. Amazon")
        print("3. Bestbuy")
        print("4. Aliexpress")
        print()

        opcion= int(input("Seleccione el numero de la opcion deseada\n"))

        if opcion == 1:
            busqueda_mercadolibre()
            envio_correo()
        elif opcion == 2:
            busqueda_amazon()
            envio_correo()
        elif opcion == 3:
            busqueda_bestbuy()
            envio_correo()
        elif opcion == 4:
            busqueda_aliexpress()
            envio_correo()


menu()



